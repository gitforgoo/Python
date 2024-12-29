#BzR
import pandas as pd
import os
from openpyxl import load_workbook, Workbook
import xlrd

def read_excel_files(folder_path, output_file):
    # 폴더 내 모든 파일 목록 가져오기
    files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') or f.endswith('.xls')]

    cell_addresses = ['B4','F23','L23', 'L24', 'Q15', 'S15']
    output_data = []

    for file in files:
        file_path = os.path.join(folder_path, file)
        print(f"Reading file: {file_path}")
        
        # 파일 확장자에 따라 적절한 엔진 선택
        if file.endswith('.xlsx'):
            # openpyxl을 사용하여 엑셀 파일 열기
            #df = pd.read_excel(file_path, engine='openpyxl')
            wb = load_workbook(file_path, data_only=True)
            sheet = wb.active
        elif file.endswith('.xls'):
            # xlrd를 사용하여 엑셀 파일 열기
            wb = xlrd.open_workbook(file_path, formatting_info=True)
            sheet = wb.sheet_by_index(0)

        row_data = []  # 파일 이름을 첫 번째 열에 추가
        for cell_address in cell_addresses:
            if file.endswith('.xlsx'):
                # openpyxl을 사용하여 수식이 있는 셀의 값을 읽기
                cell_value = sheet[cell_address].value
            elif file.endswith('.xls'):
                # xlrd를 사용하여 셀의 값을 읽기
                col_letter = cell_address[0]
                row_number = int(cell_address[1:])
                col_index = ord(col_letter) - ord('A')
                cell_value = sheet.cell_value(row_number - 1, col_index)
            
            # B4 이외의 값은 100000으로 나누고 소수점 3자리에서 반올림하여 소수점 2자리까지 가져오기
            if cell_address != 'B4' and isinstance(cell_value, (int, float)):
                cell_value = round(cell_value / 1000000, 2)

            print(f"Value at {cell_address} in file {file}: {cell_value}")
            row_data.append(cell_value)
        
        output_data.append(row_data)

    # 읽어온 값을 새로운 엑셀 파일에 저장
    save_to_excel(output_data, output_file)

    print("Job Done!")
        
def save_to_excel(data, output_file):
    output_cell_addresses = ['C4', 'D4', 'F4', 'E4', 'G4', 'H4'] 

    # 기존 엑셀 파일 열기
    if os.path.exists(output_file):
        wb = load_workbook(output_file)
    else:
        wb = Workbook()
    ws = wb.active

    # 데이터 추가
    for i, row in enumerate(data):
        for j, value in enumerate(row):
            if j < len(output_cell_addresses):
                start_cell = output_cell_addresses[j]
                start_col_letter = start_cell[0]
                start_row_number = int(start_cell[1:])
                start_col_index = ord(start_col_letter) - ord('A')
                ws.cell(row=start_row_number + i, column=start_col_index + 1, value=value)

    # 엑셀 파일 저장
    wb.save(output_file)


# 사용 예시
folder_path = 'D:/03_Company/20_고객지원파트/99_원가표/2024/01_유지보수/01_BzR'
output_file = 'D:/03_Company/20_고객지원파트/01_연간계획/2025/01_경영계획/메신저_유지보수_수주_수행내역_BzR.xlsx'
read_excel_files(folder_path, output_file)

folder_path = 'D:/03_Company/20_고객지원파트/99_원가표/2024/01_유지보수/02_메신저'
output_file = 'D:/03_Company/20_고객지원파트/01_연간계획/2025/01_경영계획/메신저_유지보수_수주_수행내역_MSG.xlsx'
read_excel_files(folder_path, output_file)